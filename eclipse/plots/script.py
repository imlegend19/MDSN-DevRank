import openpyxl

RELATIVE_PATH = "/home/imlegend19/PycharmProjects/Research - Data Mining/eclipse/analysis/"


def fill_ranks_lst(sheet):
    lst = {}

    for _ in range(2, sheet.max_row):
        lst[int(sheet.cell(_, 1).value)] = float(sheet.cell(_, 12).value)

    return lst


def fetch_wb(year):
    wb = openpyxl.load_workbook(RELATIVE_PATH + "analysis_" + str(year) + ".xlsx")

    return wb.active


an_2002 = fill_ranks_lst(fetch_wb(2002))
an_2003 = fill_ranks_lst(fetch_wb(2003))
an_2004 = fill_ranks_lst(fetch_wb(2004))
an_2005 = fill_ranks_lst(fetch_wb(2005))
an_2006 = fill_ranks_lst(fetch_wb(2006))

ass_2002 = [37, 34, 39, 18, 481]
ass_2003 = [51, 34, 39, 4502, 481]
ass_2004 = [61, 34, 39, 4502, 481]
ass_2005 = [329, 8707, 39, 4502, 481]
ass_2006 = ass_2002

for i in range(5):
    print("plot_" + str(ass_2002[i]) + " =",
          [an_2002[ass_2002[i]], an_2003[ass_2003[i]], an_2004[ass_2004[i]], an_2005[ass_2005[i]],
           an_2006[ass_2006[i]]])
