import pickle
import openpyxl

wb_obj = openpyxl.load_workbook("layer1.xlsx")

sheet_obj = wb_obj.active

layer = 1

degree = {}
ec = {}
close = {}
harm_close = {}
bet = {}
pg = {}
for i in range(2, sheet_obj.max_row+1):
    degree[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 2).value
    close[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 3).value
    harm_close[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 4).value
    bet[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 5).value
    ec[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 6).value
    pg[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 7).value

with open("degree/degree_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(degree, fp)

with open("eigenvector/eigenvector_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(ec, fp)

with open("closeness/closeness_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(close, fp)

with open("betweenness/betweenness_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(bet, fp)

with open("harmonic_closeness/harmonic_closeness_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(harm_close, fp)

with open("pagerank/pagerank_layer_" + str(layer) + ".txt", 'wb') as fp:
    pickle.dump(pg, fp)

print(pg)
print(harm_close)
