import pickle
from itertools import permutations

import openpyxl

RELATIVE_PATH = "/home/imlegend19/PycharmProjects/Research - Data Mining/eclipse/ranks_normal/definition_2/"
NO_OF_LAYERS = 3


def fill_ranks_lst(sheet):
    lst = []

    for _ in range(3, sheet.max_row):
        lst.append((int(sheet.cell(_, 1).value), int(sheet.cell(_, 2).value)))

    return lst


def fetch_wb(layer):
    wb = openpyxl.load_workbook(RELATIVE_PATH + "layer" + str(layer) + "_ranks_fc.xlsx")

    return wb.active


def calculate_influence(layer1, layer2):
    match = 0
    x = len(layer1)
    y = len(layer2)

    z = max(x, y)
    if x == z:
        for i in range(y):
            if layer1[i][1] == layer2[i][1]:
                match += 1
    else:
        for i in range(x):
            if layer1[i][1] == layer2[i][1]:
                match += 1

    return match / x


layer_list = []
for _ in range(NO_OF_LAYERS):
    layer_list.append(fill_ranks_lst(fetch_wb(_ + 1)))

pairs = list(permutations([_ for _ in range(1, NO_OF_LAYERS + 1)], 2))
for _ in pairs:
    if _[0] == _[1]:
        pairs.remove(_)

print(pairs)

influence_matrix = [[1, 0, 0],
                    [0, 1, 0],
                    [0, 0, 1]]

for i in pairs:
    print("Ongoing pair: %d, %d" % (i[0], i[1]))
    a = int(i[0])
    b = int(i[1])

    w = calculate_influence(layer_list[a - 1], layer_list[b - 1])

    influence_matrix[a - 1][b - 1] = w

with open("influence_matrix_fc.txt", 'wb') as fp:
    pickle.dump(influence_matrix, fp)
