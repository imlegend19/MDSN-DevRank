import openpyxl
from itertools import permutations

RELATIVE_PATH = "/home/imlegend19/PycharmProjects/Research - Data Mining/ranks/"
NO_OF_LAYERS = 6
TOP = [10, 20, 50, 100, 500, 1000]


def fill_ranks_lst(sheet):
    lst = []

    for _ in range(3, sheet.max_row):
        lst.append((int(sheet.cell(_, 1).value), int(sheet.cell(_, 2).value)))

    return lst


def fill_ranks_dic(sheet):
    dic = {}

    for _ in range(3, sheet.max_row):
        dic[int(sheet.cell(_, 2).value)] = int(sheet.cell(_, 1).value)

    return dic


def fetch_wb(layer):
    wb = openpyxl.load_workbook(RELATIVE_PATH + "layer" + str(layer) + "_ranks_fc.xlsx")

    return wb.active


def fetch_top(ranks, num):
    return ranks[:num]


def calculate_match(a, b, top):
    match_counter = 0

    for i in range(top):
        try:
            if a[i][1] == b[i][1]:
                match_counter += 1
        except IndexError:
            break

    return match_counter / top * 100


def calculate_shift(a, b, top, a_dic, b_dic):
    sft = 0

    for i in range(top):
        try:
            if a[i][1] != b[i][1]:
                x = a_dic[a[i][1]]
                try:
                    y = b_dic[a[i][1]]
                except KeyError:
                    y = len(b_dic.keys())
                sft += abs(x - y)
        except IndexError:
            break

    return sft / top


"""
Top 10, 20, 50, 100, 500, 1000
Mismatches
Average difference
"""

print("Fetching ranks of all layers...")

layer_dict = []
for _ in range(NO_OF_LAYERS):
    layer_dict.append(fill_ranks_dic(fetch_wb(_ + 1)))

layer_list = []
for _ in range(NO_OF_LAYERS):
    layer_list.append(fill_ranks_lst(fetch_wb(_ + 1)))

pairs = list(permutations([_ for _ in range(1, NO_OF_LAYERS + 1)], 2))
for _ in pairs:
    if _[0] == _[1]:
        pairs.remove(_)

print(pairs)

for i in pairs:
    print("Ongoing pair: %d, %d" % (i[0], i[1]))

    a, b = i[0], i[1]

    wb = openpyxl.Workbook()
    sh = wb.active

    sh.append(["Top", "Match Percent", "Average Shift"])
    sh.append(["", "", ""])

    for top in TOP:
        match_percent = calculate_match(layer_list[a - 1], layer_list[b - 1], top)
        shift = calculate_shift(layer_list[a - 1], layer_list[b - 1], top, layer_dict[a - 1], layer_dict[b - 1])

        sh.append([str(top), str(match_percent), str(shift)])

    file_name = str(a) + "-" + str(b) + "-analysis-fc.xlsx"
    wb.save(file_name)

print("Process Complete!")
