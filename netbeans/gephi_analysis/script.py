import pickle
import openpyxl

wb_obj = openpyxl.load_workbook("combined.xlsx")

sheet_obj = wb_obj.active

degree = {}
ec = {}
close = {}
harm_close = {}
bet = {}
pg = {}
for i in range(2, sheet_obj.max_row+1):
    degree[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 2)
    close[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 3)
    harm_close[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 4)
    bet[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 5)
    ec[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 6)
    pg[sheet_obj.cell(i, 1).value] = sheet_obj.cell(i, 7)

with open("degree/degree_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(degree, fp)

with open("eigenvector/eigenvector_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(ec, fp)

with open("closeness/closeness_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(close, fp)

with open("betweenness/betweenness_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(bet, fp)

with open("harmonic_closeness/harmonic_closeness_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(harm_close, fp)

with open("pagerank/pagerank_layer_combined" + ".txt", 'wb') as fp:
    pickle.dump(pg, fp)
