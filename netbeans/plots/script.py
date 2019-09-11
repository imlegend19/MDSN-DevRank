import openpyxl

RELATIVE_PATH = "/home/imlegend19/PycharmProjects/Research - Data Mining/netbeans/analysis/"


def fill_ranks_lst(sheet):
    lst = {}

    for _ in range(2, sheet.max_row):
        lst[int(sheet.cell(_, 1).value)] = float(sheet.cell(_, 12).value)

    return lst


def fetch_wb(year):
    wb = openpyxl.load_workbook(RELATIVE_PATH + "analysis_" + str(year) + ".xlsx")

    return wb.active


an_2001 = fill_ranks_lst(fetch_wb(2001))
an_2002 = fill_ranks_lst(fetch_wb(2002))
an_2003 = fill_ranks_lst(fetch_wb(2003))
an_2004 = fill_ranks_lst(fetch_wb(2004))
an_2005 = fill_ranks_lst(fetch_wb(2005))

ass = [6, 258, 285, 190, 123]

for i in ass:
    data = [an_2001[i], an_2002[i], an_2003[i], an_2004[i], an_2005[i]]
    print("plot_" + str(i) + " =", data)
