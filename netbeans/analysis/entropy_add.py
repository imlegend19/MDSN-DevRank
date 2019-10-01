import pickle

import openpyxl

wb_obj = openpyxl.load_workbook("netbeans.xlsx")
sheet_obj = wb_obj.active

wb = openpyxl.Workbook()
sheet = wb.active


def open_file(path):
    with open(path, 'rb') as fp:
        return pickle.load(fp)


ent_l1 = open_file("layer1_entropy.txt")
ent_l2_d1 = open_file("layer2_d1_entropy.txt")
ent_l2_d2 = open_file("layer2_d2_entropy.txt")
ent_l3 = open_file("layer3_entropy.txt")
ent_l4 = open_file("layer4_entropy.txt")
ent_comb = open_file("entropy_comb.txt")

for i in range(1, sheet_obj.max_row + 1):
    row = []
    for j in range(1, sheet_obj.max_column + 1):
        row.append(sheet_obj.cell(i, j).value)

    try:
        if i == 1:
            row.extend(['Ent-L1', 'Ent-L2-D1', 'Ent-L2-D2', 'Ent-L3', 'Ent-L4', 'Ent-Comb'])
        else:
            row.extend(
                [abs(ent_l1[int(row[0])]), abs(ent_l2_d1[int(row[0])]), abs(ent_l2_d2[int(row[0])]),
                 abs(ent_l3[int(row[0])]), abs(ent_l4[int(row[0])]), abs(ent_comb[int(row[0])])])

        sheet.append(row)
    except Exception:
        print(row[0])

wb.save("netbeans_updated.xlsx")
